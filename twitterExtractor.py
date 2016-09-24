import twitter
import argparse

from openpyxl import Workbook
import openpyxl.compat
from openpyxl.utils import get_column_letter
from TwitterSearch import *



try:

# -----------------------------------------------------------------
# Parameters
# ----------------------------------------------------------------
    query = ['#Zika']
    language = 'en'

# Result file
    wb = Workbook()
    file_name = "results.xlsx"
    ws = wb.active
    ws.title = "Results query "
    print 'File created'

# Twitter Search API
    tso = TwitterSearchOrder() # create a TwitterSearchOrder object
    tso.set_keywords(query) # let's define all words we would like to have a look for
    tso.set_language(language) # we want to see English tweets only
    tso.set_include_entities(False) # and don't give us all those entity information

    # Create a TwitterSearch object with the secret tokens
    ts = TwitterSearch(
        consumer_key = 'Qpbao6aMCktKEyp67cxi36Hyl',
        consumer_secret = 'LmaqPEr0nwIhcmPcBfEvDiiuJhERifot80AafC7jJmnOq4T6w0',
        access_token = '2841200026-gQSl3UCP3VMhFKpQWJNfXjY0WX846KL4idI2YPU',
        access_token_secret = 'eFpItW0liT4gmPeTaB3wJRMkc98SmReAHXNqRtdIB92XC'
     )

    inc = 1
     # Querying twitter
    for tweet in ts.search_tweets_iterable(tso):
        tweetText = '@%s tweeted: %s' % ( tweet['user']['screen_name'], tweet['text'])

        r = 'A' + str(inc)
        ws[r] = tweetText #Storing every tweet in the excel file

        inc += 1


except TwitterSearchException as e: # Handling errors
    print(e)


